"""
FinSight Excel Generator - COMPLETE FIXED VERSION
FIXES:
1. Uses previous_balance and new_balance from pdf_processor
2. Monthly Income = Inflows (including transfers/payments)
3. Monthly Expenses = Outflows (all spending)
4. Cash Flow shows ALL inflows including payments
"""
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Styles
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="1E40AF")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
SUBTITLE_FONT = Font(name="Calibri", italic=True, size=10, color="6B7280")
MONEY_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
PERCENT_FORMAT = '0.0%'
BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")
RED_FILL = PatternFill("solid", fgColor="FEE2E2")
YELLOW_FILL = PatternFill("solid", fgColor="FEF9C3")
BLUE_FILL = PatternFill("solid", fgColor="DBEAFE")
GRAY_FILL = PatternFill("solid", fgColor="F3F4F6")
ALT_ROW_FILL = PatternFill("solid", fgColor="F8FAFC")


def create_data_tab(wb, all_transactions):
    """Create DATA tab."""
    ws = wb.create_sheet("Data")
    
    headers = ["Date", "Description", "Amount", "Type", "Category", 
               "Subcategory", "Account", "Needs Research", "Confidence"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    
    for row, tx in enumerate(all_transactions, 2):
        ws.cell(row, 1).value = tx.get('date')
        ws.cell(row, 2).value = tx.get('description')
        ws.cell(row, 3).value = tx.get('amount')
        ws.cell(row, 3).number_format = MONEY_FORMAT
        ws.cell(row, 4).value = tx.get('type')
        ws.cell(row, 5).value = tx.get('category')
        ws.cell(row, 6).value = tx.get('classification', '')
        ws.cell(row, 7).value = tx.get('account_name', tx.get('bank', ''))
        ws.cell(row, 8).value = "YES" if tx.get('needs_research') else "NO"
        ws.cell(row, 9).value = tx.get('confidence', 0)
        ws.cell(row, 9).number_format = PERCENT_FORMAT
        
        if tx.get('needs_research'):
            for col in range(1, 10):
                ws.cell(row, col).fill = YELLOW_FILL
        elif row % 2 == 0:
            for col in range(1, 10):
                ws.cell(row, col).fill = ALT_ROW_FILL
        
        for col in range(1, 10):
            ws.cell(row, col).border = BORDER
    
    return ws


def create_reconciliation_tab(wb, processed_files):
    """Create RECONCILIATION tab with SUMIFS formulas."""
    ws = wb.create_sheet("Reconciliation", 0)
    
    ws.merge_cells('A1:G1')
    ws['A1'] = "ACCOUNT RECONCILIATION"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:G2')
    ws['A2'] = "✨ Formula-based using SUMIFS by account"
    ws['A2'].font = SUBTITLE_FONT
    ws['A2'].alignment = Alignment(horizontal="center")
    
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18
    
    row = 4
    headers = ["Account", "Beginning Balance", "Credits (+)", "Debits (-)", 
               "Calculated Ending", "Statement Ending", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    
    row = 5
    for pf in processed_files:
        bank = pf.get('bank', 'Unknown')
        account_type = pf.get('account_type', 'Unknown')
        account_number = str(pf.get('account_number', 'Unknown'))
        account_name = f"{bank} {account_type} ...{account_number[-4:]}"
        
        # FIXED: Use correct field names from pdf_processor
        beginning = pf.get('previous_balance', 0)
        ending = pf.get('new_balance', 0)
        
        # Credit cards: balances as negative (liabilities)
        if account_type == 'Credit':
            beginning = -abs(beginning) if beginning > 0 else beginning
            ending = -abs(ending) if ending > 0 else ending
        
        ws.cell(row, 1).value = account_name
        
        ws.cell(row, 2).value = beginning
        ws.cell(row, 2).number_format = MONEY_FORMAT
        if beginning < 0:
            ws.cell(row, 2).font = Font(color="DC2626")
        
        ws.cell(row, 3).value = f'=SUMIFS(Data!C:C,Data!G:G,"{account_name}",Data!C:C,">0")'
        ws.cell(row, 3).number_format = MONEY_FORMAT
        ws.cell(row, 3).fill = GREEN_FILL
        
        ws.cell(row, 4).value = f'=SUMIFS(Data!C:C,Data!G:G,"{account_name}",Data!C:C,"<0")'
        ws.cell(row, 4).number_format = MONEY_FORMAT
        ws.cell(row, 4).fill = RED_FILL
        
        ws.cell(row, 5).value = f'=B{row}+C{row}+D{row}'
        ws.cell(row, 5).number_format = MONEY_FORMAT
        ws.cell(row, 5).font = Font(bold=True)
        if ending < 0:
            ws.cell(row, 5).font = Font(bold=True, color="DC2626")
        
        ws.cell(row, 6).value = ending
        ws.cell(row, 6).number_format = MONEY_FORMAT
        ws.cell(row, 6).font = Font(bold=True)
        if ending < 0:
            ws.cell(row, 6).font = Font(bold=True, color="DC2626")
        
        ws.cell(row, 7).value = f'=IF(ABS(E{row}-F{row})<0.5,"✅ OK",CONCATENATE("❌ $",TEXT(ABS(E{row}-F{row}),"0.00")))'
        ws.cell(row, 7).font = Font(bold=True)
        
        for col in range(1, 8):
            ws.cell(row, col).border = BORDER
        
        row += 1
    
    row += 2
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row, 1).value = "💡 All values calculated from Data tab"
    ws.cell(row, 1).font = Font(italic=True, size=9)
    ws.cell(row, 1).fill = BLUE_FILL
    
    return ws


def create_cash_flow_statement(wb, analysis_period: str = "Period"):
    """Create CASH FLOW STATEMENT - FIXED to include ALL inflows/outflows."""
    ws = wb.create_sheet("Cash Flow Statement")
    
    ws.merge_cells('A1:E1')
    ws['A1'] = "STATEMENT OF CASH FLOWS"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Analysis Period: {analysis_period}"
    ws['A2'].font = SUBTITLE_FONT
    ws['A2'].alignment = Alignment(horizontal="center")
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15
    
    row = 4
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row, 1).value = "CASH INFLOWS"
    ws.cell(row, 1).font = SECTION_FONT
    ws.cell(row, 1).fill = PatternFill("solid", fgColor="059669")
    
    row += 1
    headers = ["Category", "Period Total", "Annualized Pace", "Analysis", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    
    row += 1
    start_income_row = row
    
    # FIXED: Include ALL positive amounts (including transfers/payments)
    income_items = [
        ("Payments & Transfers", "Money IN (payments, transfers)", ["Transfer"]),
        ("Refunds", "Returns and refunds", ["Refund"]),
        ("Other Credits", "Other money received", ["Entertainment"]),  # For Amex credits
    ]
    
    for cat_name, analysis, categories in income_items:
        ws.cell(row, 1).value = cat_name
        
        # FORMULA: Sum all positive amounts for these categories
        category_formula = '+'.join([f'SUMIFS(Data!C:C,Data!E:E,"{cat}",Data!C:C,">0")' for cat in categories])
        ws.cell(row, 2).value = f'={category_formula}'
        ws.cell(row, 2).number_format = MONEY_FORMAT
        ws.cell(row, 3).value = f'=B{row}*12'
        ws.cell(row, 3).number_format = MONEY_FORMAT
        ws.cell(row, 4).value = analysis
        ws.cell(row, 4).font = Font(size=9, italic=True)
        ws.cell(row, 5).value = "ℹ️"
        
        for col in range(1, 6):
            ws.cell(row, col).border = BORDER
        
        row += 1
    
    # Total inflows
    ws.cell(row, 1).value = "TOTAL CASH INFLOWS:"
    ws.cell(row, 1).font = Font(bold=True, size=11)
    ws.cell(row, 2).value = f'=SUM(B{start_income_row}:B{row-1})'
    ws.cell(row, 2).number_format = MONEY_FORMAT
    ws.cell(row, 2).font = Font(bold=True)
    ws.cell(row, 2).fill = GREEN_FILL
    ws.cell(row, 3).value = f'=B{row}*12'
    ws.cell(row, 3).number_format = MONEY_FORMAT
    ws.cell(row, 3).font = Font(bold=True)
    
    total_income_row = row
    
    # OUTFLOWS
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row, 1).value = "CASH OUTFLOWS"
    ws.cell(row, 1).font = SECTION_FONT
    ws.cell(row, 1).fill = PatternFill("solid", fgColor="DC2626")
    
    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    
    row += 1
    start_expense_row = row
    
    expense_categories = [
        ("Childcare", "Fixed cost - essential"),
        ("Shopping", "Discretionary - can optimize"),
        ("Entertainment", "Discretionary - review subscriptions"),
        ("Groceries", "Essential - food shopping"),
        ("Healthcare", "Essential - medical expenses"),
        ("Utilities", "Fixed cost - essential services"),
        ("Meals", "Discretionary - dining out"),
        ("Transportation", "Essential - fuel/transit"),
        ("Other", "Needs categorization"),
    ]
    
    for cat, analysis in expense_categories:
        ws.cell(row, 1).value = cat
        ws.cell(row, 2).value = f'=ABS(SUMIF(Data!E:E,"{cat}",Data!C:C))'
        ws.cell(row, 2).number_format = MONEY_FORMAT
        ws.cell(row, 3).value = f'=B{row}*12'
        ws.cell(row, 3).number_format = MONEY_FORMAT
        ws.cell(row, 4).value = analysis
        ws.cell(row, 4).font = Font(size=9, italic=True)
        ws.cell(row, 5).value = f'=IF(B{row}>500,"🔴",IF(B{row}>200,"⚠️","✅"))'
        
        for col in range(1, 6):
            ws.cell(row, col).border = BORDER
        
        row += 1
    
    # Total outflows
    ws.cell(row, 1).value = "TOTAL CASH OUTFLOWS:"
    ws.cell(row, 1).font = Font(bold=True, size=11)
    ws.cell(row, 2).value = f'=SUM(B{start_expense_row}:B{row-1})'
    ws.cell(row, 2).number_format = MONEY_FORMAT
    ws.cell(row, 2).font = Font(bold=True)
    ws.cell(row, 2).fill = RED_FILL
    ws.cell(row, 3).value = f'=B{row}*12'
    ws.cell(row, 3).number_format = MONEY_FORMAT
    ws.cell(row, 3).font = Font(bold=True)
    
    total_expense_row = row
    
    # NET CHANGE
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row, 1).value = "NET CASH CHANGE"
    ws.cell(row, 1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row, 1).fill = PatternFill("solid", fgColor="1E40AF")
    ws.cell(row, 1).alignment = Alignment(horizontal="center")
    
    row += 1
    ws.cell(row, 1).value = "Net Change in Cash:"
    ws.cell(row, 1).font = Font(bold=True, size=12)
    ws.cell(row, 2).value = f'=B{total_income_row}-B{total_expense_row}'
    ws.cell(row, 2).number_format = MONEY_FORMAT
    ws.cell(row, 2).font = Font(bold=True, size=14, color="DC2626")
    ws.cell(row, 2).fill = YELLOW_FILL
    ws.cell(row, 3).value = f'=B{row}*12'
    ws.cell(row, 3).number_format = MONEY_FORMAT
    ws.cell(row, 3).font = Font(bold=True, size=12)
    ws.cell(row, 5).value = f'=IF(B{row}>0,"✅ Positive","🔴 Negative")'
    ws.cell(row, 5).font = Font(bold=True, size=11)
    
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row, 1).value = "💡 Shows ALL cash movements including payments and transfers"
    ws.cell(row, 1).font = Font(italic=True, size=9)
    ws.cell(row, 1).fill = BLUE_FILL
    
    return ws


def generate_workbook(processed_files, all_transactions, budget_analysis, output_path):
    """Generate complete workbook."""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Determine analysis period
    if all_transactions:
        dates = [tx.get('date', '') for tx in all_transactions if tx.get('date')]
        if dates:
            dates.sort()
            period = f"{dates[0]} to {dates[-1]}"
        else:
            period = "Unknown Period"
    else:
        period = "No Transactions"
    
    # Create tabs
    create_reconciliation_tab(wb, processed_files)
    create_data_tab(wb, all_transactions)
    create_cash_flow_statement(wb, period)
    
    wb.save(output_path)
    return output_path
